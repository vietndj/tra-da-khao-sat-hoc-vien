import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_survey_excel():
    wb = openpyxl.Workbook()
    
    # Sheet 1: Raw Data
    ws_data = wb.active
    ws_data.title = "Dữ Liệu Khảo Sát"
    
    # Colors
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    
    headers = [
        "Mã Phản Hồi",
        "Thời Gian Gửi",
        "Khóa Học",
        "Họ Và Tên",
        "Số Điện Thoại / Zalo",
        "Nghề Nghiệp / Dự Án",
        "Link Kênh / Profile",
        "Nguồn Biết Đến",
        "Video Ấn Tượng Nhất",
        "Điều Tâm Đắc Nhất",
        "Góp Ý Vận Hành & Lớp Học",
        "Điểm Đánh Giá (1-10)",
        "Nhu Cầu Khóa Nâng Cao",
        "Kênh Liên Hệ Ưu Tiên",
        "Lời Nhắn Nhủ Riêng"
    ]
    
    ws_data.append(headers)
    
    # Format header row
    for col_num in range(1, len(headers) + 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws_data.row_dimensions[1].height = 32
    
    # Sample Row from actual feedback
    sample_rows = [
        [
            "KS-2026-001",
            "24/08/2026 09:30:15",
            "Offline Thực Chiến K12 - Hà Nội",
            "Nguyễn Thu Trang",
            "0912345678",
            "Quản lý pháp chế MEDLATEC (Chuẩn bị mở cty luật)",
            "https://facebook.com/thutrang.law",
            "Video TikTok / Reel (Quyết định sau 20s xem video)",
            "Video anh Việt nói về băm nhỏ phân cảnh và tư duy làm video AI đa ngành",
            "Tư duy rất mới về ứng dụng AI, kỹ năng quay dựng và cách làm marketing đa ngành, định hướng rõ ràng",
            "- Cần kiểm soát danh sách check-in chặt chẽ hơn\n- Gửi link khảo sát ngắn trước buổi học\n- Thăm hỏi học viên vắng/về sớm\n- Giới thiệu gói nâng cao 10p cuối kèm ưu đãi tại lớp\n- Tư vấn rõ ưu đãi giảm 10% khi đăng ký nhóm 2 người",
            10,
            "AI Chuyên Sâu, Xây dựng thương hiệu cá nhân Luật Sư",
            "Zalo",
            "Giá trị Thầy mang lại là điều vô giá đối với Em ạ! Chúc Thầy luôn giữ ngọn lửa nhiệt huyết."
        ],
        [
            "KS-2026-002",
            "24/08/2026 10:15:40",
            "Offline Thực Chiến K12 - Hà Nội",
            "Trần Hoàng Nam",
            "0988776655",
            "Kinh doanh thời trang & phụ kiện Hải Phòng",
            "https://tiktok.com/@nam.storehp",
            "YouTube Shorts",
            "Video phân tích kịch bản video bán hàng thời trang 3 tầng",
            "Biết cách setup 2 máy quay và ánh sáng chuyên nghiệp, không còn sợ đứng trước ống kính",
            "Khóa học rất thực chiến, mong có thêm buổi thực hành ngoài trời",
            9,
            "Video Ads Chuyển Đổi Cao, Automation Video",
            "Zalo",
            "Cảm ơn anh Việt rất nhiều, nhờ anh mà em tự tin bấm máy quay ngay ngày hôm sau!"
        ]
    ]
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    for row in sample_rows:
        ws_data.append(row)
        
    for r in range(2, len(sample_rows) + 2):
        ws_data.row_dimensions[r].height = 48
        for c in range(1, len(headers) + 1):
            cell = ws_data.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            if c in [1, 2, 5, 12, 14]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if c == 12:
                cell.font = Font(name="Arial", size=11, bold=True, color="059669")
                
    # Auto-adjust column widths
    column_widths = {
        1: 16, 2: 20, 3: 28, 4: 22, 5: 18, 6: 30, 7: 30,
        8: 25, 9: 35, 10: 38, 11: 45, 12: 15, 13: 32, 14: 18, 15: 40
    }
    for col_idx, width in column_widths.items():
        ws_data.column_dimensions[get_column_letter(col_idx)].width = width
        
    # Sheet 2: Dashboard Overview
    ws_dash = wb.create_sheet(title="Dashboard Tổng Quan")
    
    dash_title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    dash_title_font = Font(name="Arial", size=14, bold=True, color="F8FAFC")
    
    ws_dash.merge_cells("A1:F1")
    title_cell = ws_dash["A1"]
    title_cell.value = "📊 BÁO CÁO TỔNG HỢP KHẢO SÁT & INSIGHT HỌC VIÊN - VIETMAC"
    title_cell.fill = dash_title_fill
    title_cell.font = dash_title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 36
    
    # KPI Metric Cards
    metrics = [
        ("Tổng Số Phản Hồi", "=COUNTA('Dữ Liệu Khảo Sát'!A2:A1000)", "A3:B4", "0284C7"),
        ("Điểm Hài Lòng Trung Bình (NPS)", "=AVERAGE('Dữ Liệu Khảo Sát'!L2:L1000)", "C3:D4", "059669"),
        ("Tỷ Lệ Quan Tâm Nâng Cao", "=COUNTIF('Dữ Liệu Khảo Sát'!M2:M1000, \"*AI*\")/COUNTA('Dữ Liệu Khảo Sát'!A2:A1000)", "E3:F4", "D97706")
    ]
    
    ws_dash.row_dimensions[3].height = 20
    ws_dash.row_dimensions[4].height = 28
    
    for title, formula, cell_range, color_hex in metrics:
        start_col = cell_range.split(":")[0][0]
        start_row = int(cell_range.split(":")[0][1:])
        end_col = cell_range.split(":")[1][0]
        end_row = int(cell_range.split(":")[1][1:])
        
        ws_dash.merge_cells(f"{start_col}{start_row}:{end_col}{start_row}")
        lbl_cell = ws_dash[f"{start_col}{start_row}"]
        lbl_cell.value = title
        lbl_cell.font = Font(name="Arial", size=9, bold=True, color="64748B")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws_dash.merge_cells(f"{start_col}{end_row}:{end_col}{end_row}")
        val_cell = ws_dash[f"{start_col}{end_row}"]
        val_cell.value = formula
        val_cell.font = Font(name="Arial", size=16, bold=True, color=color_hex)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Border
        for r in range(start_row, end_row + 1):
            for c_letter in [start_col, end_col]:
                col_idx = openpyxl.utils.column_index_from_string(c_letter)
                ws_dash.cell(row=r, column=col_idx).border = thin_border
                
    for col in range(1, 8):
        ws_dash.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 3: CRM Follow-up View
    ws_crm = wb.create_sheet(title="CRM Kết Nối & Tặng Quà")
    crm_headers = ["STT", "Họ Tên Học Viên", "Nghề Nghiệp / Dự Án", "Link Kênh / Profile", "Zalo / SĐT", "Nhu Cầu Độc Quyền", "Đã Kết Nối / Tặng Quà"]
    ws_crm.append(crm_headers)
    
    for col_num in range(1, len(crm_headers) + 1):
        cell = ws_crm.cell(row=1, column=col_num)
        cell.fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_crm.row_dimensions[1].height = 28
    
    crm_sample = [
        [1, "Nguyễn Thu Trang", "Quản lý pháp chế MEDLATEC (Mở cty luật)", "https://facebook.com/thutrang.law", "0912345678", "Gói prompt AI kịch bản luật sư + template branding", "Đã add Zalo & gửi quà"],
        [2, "Trần Hoàng Nam", "Kinh doanh thời trang & phụ kiện", "https://tiktok.com/@nam.storehp", "0988776655", "Checklist setup 2 máy quay + sound effect bán hàng", "Đã follow kênh TikTok"]
    ]
    for row in crm_sample:
        ws_crm.append(row)
    
    for r in range(2, len(crm_sample) + 2):
        ws_crm.row_dimensions[r].height = 36
        for c in range(1, len(crm_headers) + 1):
            cell = ws_crm.cell(row=r, column=c)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            if c in [1, 5, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    crm_widths = {1: 8, 2: 24, 3: 35, 4: 32, 5: 18, 6: 45, 7: 25}
    for col_idx, width in crm_widths.items():
        ws_crm.column_dimensions[get_column_letter(col_idx)].width = width
        
    output_filename = "Mau_Bang_Khao_Sat_Hoc_Vien_VietMac.xlsx"
    wb.save(output_filename)
    print(f"Successfully generated {output_filename}")

if __name__ == "__main__":
    create_survey_excel()
