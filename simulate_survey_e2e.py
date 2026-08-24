import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_survey_excel():
    excel_file = "/Users/vietmac/Documents/CODE/tra-da-khao-sat-hoc-vien/Mau_Bang_Khao_Sat_Hoc_Vien_VietMac.xlsx"
    wb = openpyxl.Workbook()
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # -------------------------------------------------------------
    # TAB 1: DỮ LIỆU KHẢO SÁT & ẢNH KỶ NIỆM
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Dữ Liệu Khảo Sát & Ảnh"
    ws1.views.sheetView[0].showGridLines = True
    
    headers1 = [
        "Mã Phản Hồi",
        "Thời Gian Gửi",
        "Khóa Học",
        "Họ Và Tên",
        "Số Zalo Hay Dùng",
        "Mảng Kinh Doanh & Định Hướng AI",
        "Danh Sách Link Kênh / Video / FB",
        "Hành Trình Biết Đến & Xem Video",
        "Góp Ý Thẳng Thắn & Lời Nhắn",
        "Số Lượng Ảnh",
        "Link Drive & Tên File Ảnh Kỷ Niệm"
    ]
    ws1.append(headers1)
    
    # Header format
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    
    for col_num in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws1.row_dimensions[1].height = 32
    
    # Sample Submissions with 2 photos
    submissions = [
        {
            "id": "KS-2026-0001",
            "time": "24/08/2026 08:15:20",
            "course": "Lớp Offline Thực Chiến - Hà Nội",
            "name": "Nguyễn Thu Trang",
            "phone": "0912345678",
            "profession": "Luật sư pháp chế (Chuẩn bị mở văn phòng luật riêng). Muốn làm dạng video case 45s như kênh tiktok.com/@luatsucuocsong. Cần AI bóc tách kịch bản và setup 2 góc máy.",
            "links": "https://facebook.com/thutrang.law\nhttps://tiktok.com/@thutrang_legal",
            "journey": "Lướt thấy video anh Việt chia sẻ về băm phân cảnh 3 tầng và góc nhìn AI thực chiến. Xem đúng 20 giây là thấy cách nói quá mộc mạc, không đao to búa lớn nên nhắn tin đăng ký và chuyển khoản luôn.",
            "feedback": "• Khâu tư vấn: Các bạn nên tư vấn rõ ưu đãi giảm 10% khi đăng ký nhóm từ đầu để học viên rủ thêm bạn.\n• Lớp học: Nhịp dạy rất thực chiến, tâm đắc nhất là tự tin cầm máy và tư duy băm nhỏ phân cảnh.\n• Lời nhắn: Cảm ơn anh nhiều, hôm nào rảnh em mời anh ly cafe nhé!",
            "photoCount": "2 ảnh",
            "photoDetails": "📁 Link Drive: https://drive.google.com/drive/folders/1aBcDeFgHiJkLmNoPqRsTuVwXyZ_NguyenThuTrang\n1. lop_hoc_thuc_hanh_quay_2_may.jpg (185 KB)\n2. anh_chup_chung_anh_viet.jpg (210 KB)"
        },
        {
            "id": "KS-2026-0002",
            "time": "24/08/2026 08:20:45",
            "course": "Lớp Offline Thực Chiến - Hà Nội",
            "name": "Trần Quốc Huy",
            "phone": "0987654321",
            "profession": "Chủ thương hiệu thời trang nam tại Hải Phòng. Muốn làm kênh chia sẻ cách phối đồ và video ads chuyển đổi cao.",
            "links": "https://tiktok.com/@huytran.menswear\nhttps://youtube.com/@huytranstyle",
            "journey": "Thấy video Facebook Reels anh Việt bóc tách kịch bản 3 tầng bán quần áo. Thấy chuẩn bài quá nên tìm kênh xem thêm 3 video nữa rồi bấm đăng ký luôn.",
            "feedback": "• 2 ngày học rất đã, vỡ ra cách setup ánh sáng và cách bóc tách video đối thủ.\n• Góp ý: Khâu check-in buổi sáng nên gửi link khảo sát trước để anh nắm nhu cầu từng người sớm hơn.\n• Nhắn nhủ: Về Hải Phòng em bấm máy quay ngay loạt video đầu tiên gửi anh xem nhé!",
            "photoCount": "2 ảnh",
            "photoDetails": "📁 Link Drive: https://drive.google.com/drive/folders/1xYzAbCdEfGhIjKlMnOpQrStUvWx_TranQuocHuy\n1. khong_khi_thuc_hanh_setup_den.jpg (192 KB)\n2. chup_ky_niem_ca_lop_ha_noi.jpg (245 KB)"
        }
    ]
    
    for row_idx, sub in enumerate(submissions, start=2):
        row_values = [
            sub["id"],
            sub["time"],
            sub["course"],
            sub["name"],
            sub["phone"],
            sub["profession"],
            sub["links"],
            sub["journey"],
            sub["feedback"],
            sub["photoCount"],
            sub["photoDetails"]
        ]
        ws1.append(row_values)
        ws1.row_dimensions[row_idx].height = 95
        
        for col_idx in range(1, len(row_values) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            
            if col_idx in [1, 2, 5, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx == 4:
                cell.font = Font(name="Arial", size=10, bold=True, color="1E3A8A")
            if col_idx == 10:
                cell.font = Font(name="Arial", size=10, bold=True, color="059669")
                cell.fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
            if col_idx == 11:
                cell.font = Font(name="Arial", size=9, color="2563EB")

    # Set column widths
    col_widths1 = {
        "A": 15, "B": 20, "C": 25, "D": 20, "E": 16,
        "F": 38, "G": 32, "H": 38, "I": 42, "J": 14, "K": 45
    }
    for col_letter, width in col_widths1.items():
        ws1.column_dimensions[col_letter].width = width

    # -------------------------------------------------------------
    # TAB 2: CRM DANH BẠ HỌC VIÊN & LINK KÊNH
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="CRM Danh Bạ & Kênh")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["STT", "Họ Và Tên", "Số Zalo", "Mảng Kinh Doanh", "Link Kênh / Video Đã Gửi", "Số Lượng Ảnh", "Trạng Thái Theo Dõi"]
    ws2.append(headers2)
    
    for c in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws2.row_dimensions[1].height = 30

    crm_data = [
        [1, "Nguyễn Thu Trang", "0912345678", "Luật sư pháp chế (Mở văn phòng)", "https://facebook.com/thutrang.law\nhttps://tiktok.com/@thutrang_legal", "2 ảnh", "Đã lưu Zalo, đã bấm follow kênh"],
        [2, "Trần Quốc Huy", "0987654321", "Thời trang nam Hải Phòng", "https://tiktok.com/@huytran.menswear\nhttps://youtube.com/@huytranstyle", "2 ảnh", "Đã lưu Zalo, đang chờ video đầu tay"]
    ]

    for r_idx, crm in enumerate(crm_data, start=2):
        ws2.append(crm)
        ws2.row_dimensions[r_idx].height = 45
        for col_idx in range(1, len(crm) + 1):
            cell = ws2.cell(row=r_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border
            if col_idx in [1, 3, 6, 7]:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx == 7:
                cell.font = Font(name="Arial", size=9, bold=True, color="059669")

    col_widths2 = {"A": 8, "B": 22, "C": 16, "D": 32, "E": 36, "F": 14, "G": 32}
    for col_letter, width in col_widths2.items():
        ws2.column_dimensions[col_letter].width = width

    # Save
    wb.save(excel_file)
    print(f"✅ Đã tạo thành công file Excel chuẩn 2 tab tại: {excel_file}")

if __name__ == "__main__":
    create_survey_excel()
